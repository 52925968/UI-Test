import os
import threading
import time

import cv2
import numpy
import openpyxl
from PIL import ImageGrab
from PIL import Image
from openpyxl.styles import PatternFill

from app_driver import App
from my_thread import MyThread
from web_driver import WEB_PAGE
import pyautogui as ui
ui.FAILSAFE = False

def elem_tuple(elem):
    el = elem.split(',')
    elem = tuple(el)
    return elem

def assert_return(app_return, expect):
    print('----------------')
    print('app_return', app_return, type(app_return))
    print('expect', expect, type(expect))
    print('----------------')
    if expect == app_return:
        return 'pass'
    elif (app_return is not None) and expect == '!None' :
        return 'pass'
    elif expect is None:
        return ''
    else:
        return 'fail'

def assert_contaion(app_return, expect):
    print('----------------')
    print('app_return', app_return, type(app_return))
    print('expect', expect, type(expect))
    print('----------------')
    for ele in app_return:
        if expect not in ele:
            return "fail"
    return "pass"

def appkey(apk,key,elem):

    if key == 'login':
        elem = elem_tuple(elem)
        return apk.login(*elem)

    if key == 'find':
        # el = elem.split(',')
        elem = elem_tuple(elem)
        return apk.find(elem)

    elif key == 'finds':
        # el = elem.split(',')
        # elem = tuple(el)
        elem = elem_tuple(elem)
        return apk.finds(elem)

    elif key == 'click':
        # el=elem.split(',')
        # elem=tuple(el)
        elem = elem_tuple(elem)
        print(elem)
        return apk.click(elem)

    elif key == 'click_word':
        print(elem)
        return apk.click_word(elem)

    elif key == 'clear':
        # el=elem.split(',')
        # elem=tuple(el)
        elem = elem_tuple(elem)
        return apk.clear(elem)


    elif key == 'clear_delete':
        # el=elem.split(',')
        # elem=tuple(el)
        elem = elem_tuple(elem)
        return apk.clear_delete(elem)

    elif key == 'input':
        # el=elem.split(',')
        # elems=tuple(el)
        elems = elem_tuple(elem)
        # print('elem',elems)
        elem1 = elems[:-1]
        word = elems[-1]
        print('elem:', elem1, 'word:', word)
        return apk.input(elem1, word)

    elif key == 'scroll':
        return apk.scroll(elem)

    elif key == 'back':
        elem = int(elem)
        return apk.back(elem)

    elif key == 'get_text':
        # el=elem.split(',')
        # elem=tuple(el)
        elem = elem_tuple(elem)
        print(elem)
        return apk.get_text(elem)

    elif key == 'get_texts':
        # el=elem.split(',')
        # elem=tuple(el)
        elem = elem_tuple(elem)
        return apk.get_texts(elem)
    # elif key=='click_word':
    #     return apk.click_word(elem)
    elif key == 'show_word':
        return apk.show_word(elem)

    elif key == 'show':
        # el=elem.split(',')
        # elem=tuple(el)
        elem = elem_tuple(elem)
        return apk.show(elem)

    elif key == 'show_word':
        print(elem)
        return apk.show_word(elem)


    elif key=='count':
        # el=elem.split(',')
        # elem=tuple(el)
        elem = elem_tuple(elem)
        return apk.count(elem)

    elif key == 'toast':
        # el=elem.split(',')
        # elem=tuple(el)
        return apk.toast(elem)

    elif key == 'handle':
        # el=elem.split(',')
        # elem=tuple(el)
        apk._driver.switch_to.window(apk._driver.window_handles[-1])
        return

    elif key == "pyautogui":
        time.sleep(1)
        print("*****************************这里是路径啊************************", elem)
        file_dir = ''.join(elem)
        time.sleep(1)
        ui.write(file_dir)
        time.sleep(0.5)
        ui.press('enter', presses=2)
        time.sleep(1)

    elif key == "select_box":
        elems = elem_tuple(elem)
        elem1 = elems[:-1]
        word = elems[-1]
        return apk.select_box(elem1, word)

    elif key == "get_attribute":
        elem = elem_tuple(elem)
        print(elem)
        return apk.get_attribute(elem)

    else:
        print('该关键字未定义', key)


def get_excel(file_path, save_path, WEB):

    book = openpyxl.load_workbook(file_path)

    table = book["Sheet1"]
    max_row = table.max_row
    print(f"Excel中的行数为：{max_row}")

    url = table.cell(1, 2).value
    print('url:', url)
    if url is not None:
        WEB.openurl(url)

    ex_title = None

    #可以设置不执行的行段
    cancel_case_start = table.cell(1, 3).value
    cancel_case_end = table.cell(1, 4).value

    res_book = openpyxl.Workbook()
    res_table = res_book.active

    for i in range(1, max_row + 1):
        no = table.cell(i, 1).value
        if no == 1:
            case_no = i
    ex_no = 0
    pass_count = 0
    fail_count = 0
    for i in range(case_no, max_row + 1):

        #如果设置了某些行不需要执行，则直接跳过
        if cancel_case_start!=None and i > cancel_case_start-1:
            if cancel_case_end == None:
                break
            elif i < cancel_case_end:
                continue

        ex_title1 = table.cell(i, 2)  # 用例名称
        title = ex_title1.value
        # print(title,type(title))
        if title is not None:
            ex_no = ex_no+1
            ex_title = str(title)
            ex_step = 1
        else:
            ex_step = ex_step+1

        ex_step_name = table.cell(i, 4)  # 步骤名称
        ex_step_name = ex_step_name.value
        print(ex_step_name)
        if ex_title == '':
            break

        ex_key = table.cell(i, 5)  #操作
        ex_key = ex_key.value

        ex_elem = table.cell(i, 6)  #元素
        ex_elem =ex_elem.value

        app_return = appkey(WEB, ex_key, ex_elem)
        time.sleep(0.5)

        ex_show = table.cell(i, 7)  #预期
        ex_show = ex_show.value


        res = assert_return(app_return, ex_show)

        print(ex_no, ex_title, ex_step, ex_step_name, ex_key, ex_elem, res)
        #将结果写入Excel中
        res_list = [ex_no, ex_title, ex_step, ex_step_name, ex_key, res, ex_show, app_return]
        fill = PatternFill("solid", fgColor="33ff00")
        fill1 = PatternFill("solid", fgColor="ff0000")

        for col, element in enumerate(res_list):
            if i == 3:
                res_table.cell(i, col + 1).value = element
                case_id_last = res_table.cell(i, 1).value
            if col == 0:
                case_id = element
            if i > 3 and case_id_last == case_id and col < 2:
                continue

            res_table.cell(i, col + 1).value = element
            # 根据结果是否通过设置不同的颜色并统计结果
            if col == 5 and element == 'fail':
                res_table.cell(i, col + 1).fill = fill1
                fail_count = fail_count + 1
            if col == 5 and element == 'pass':
                res_table.cell(i, col + 1).fill = fill
                pass_count = pass_count + 1
            # 定位元素
            res_table.cell(i, 9).value = ex_elem
        case_id_last = case_id

    res_table.cell(1, 3).value = '总计：' + str(i - 2)
    res_table.cell(1, 4).value = '通过：' + str(pass_count)
    res_table.cell(1, 4).fill = fill
    res_table.cell(1, 5).value = '失败：' + str(fail_count)
    res_table.cell(1, 5).fill = fill1
    title = (
        '序号', '用例名称', '步骤', '步骤名称', '操作', '校验值是否通过', '预期结果', '实际返回值', '元素')
    lin1 = 1
    for tt in title:
        res_table.cell(2, lin1).value = tt
        lin1 = lin1 + 1

    print('ok')
    res_book.save(save_path)

def clear_dir(path):
    """创建或者清空目录"""
    if not os.path.isdir(path):
        os.mkdir(path)
    else:
        [os.remove(os.path.join(path, file_path)) for file_path in os.listdir(path)]


if __name__ == "__main__":

    sys_time = time.strftime("%Y%m%d%H%M%S")
    file_path = r'C:\Users\admin\Desktop\ui_case-S506程控充电.xlsx'
    dir_path = r'C:\Users\admin\Desktop'
    save_name = 'res_' + sys_time + ".xlsx"
    save_path = os.path.join(dir_path, save_name)  #结果文件

    record_path = r'C:\Users\admin\Desktop\record.avi'  #视频文件

    # WEB = WEB_PAGE()
    # WEB.start()
    #
    app = App()
    app.start()

    #开启录制视频的线程
    # mythread = MyThread(record_path)
    # mythread.start()

    get_excel(file_path, save_path, app)
    time.sleep(10)


    # WEB.quit()
    #
    app.quit()



    #关闭录制视频的线程
    # mythread.stop()



